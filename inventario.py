import tkinter as tk
from tkinter import messagebox, ttk
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- CLASES DEL INVENTARIO ---
class Producto:
    def __init__(self, nombre, color, talla, stock):
        self.nombre = nombre
        self.color = color
        self.talla = talla
        self.stock = stock

    def to_dict(self):
        return {
            "nombre": self.nombre,
            "color": self.color,
            "talla": self.talla,
            "stock": self.stock
        }


class Inventario:
    def __init__(self, archivo="inventario.json"):
        self.archivo = archivo
        self.productos = self.cargar_datos()

    def guardar_datos(self):
        data = [p.to_dict() for p in self.productos]
        with open(self.archivo, "w") as f:
            json.dump(data, f, indent=4)

    def cargar_datos(self):
        if os.path.exists(self.archivo):
            try:
                with open(self.archivo, "r") as f:
                    data = json.load(f)
                    productos = []
                    for info in data:
                        if isinstance(info, dict):
                            productos.append(Producto(
                                info.get("nombre", ""),
                                info.get("color", "N/A"),
                                info.get("talla", ""),
                                info.get("stock", 0)
                            ))
                    return productos
            except json.JSONDecodeError:
                return []
        return []

    def agregar_producto(self, nombre, color, talla, stock):
        for p in self.productos:
            if (p.nombre.lower() == nombre.lower() and
                p.color.lower() == color.lower() and
                p.talla.lower() == talla.lower()):
                return False  
        self.productos.append(Producto(nombre, color, talla, stock))
        self.guardar_datos()
        return True

    def eliminar_producto(self, nombre, color, talla):
        for p in self.productos:
            if p.nombre == nombre and p.color == color and p.talla == talla:
                self.productos.remove(p)
                self.guardar_datos()
                return True
        return False

    def actualizar_producto(self, nombre, color, talla, stock):
        for p in self.productos:
            if p.nombre == nombre and p.color == color and p.talla == talla:
                p.stock = stock
                self.guardar_datos()
                return True
        return False


# --- INTERFAZ GRÁFICA ---
class InventarioApp:
    def __init__(self, root):
        self.inventario = Inventario()

        root.title("📦 Inventario de Tienda")
        root.geometry("1100x650")
        root.configure(bg="#f4f6f9")

        # Hacer que la ventana sea redimensionable
        root.rowconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)

        # --- Estilos globales ---
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Treeview",
                        font=("Segoe UI", 11),
                        rowheight=35,
                        background="white",
                        fieldbackground="white")

        style.configure("Treeview.Heading",
                        font=("Segoe UI", 12, "bold"),
                        background="#0078D7",
                        foreground="white")

        style.map("Treeview",
                  background=[("selected", "#0078D7")],
                  foreground=[("selected", "white")])

        # --- CONTENEDOR PRINCIPAL ---
        main_frame = tk.Frame(root, bg="#f4f6f9")
        main_frame.grid(row=0, column=0, sticky="nsew")

        main_frame.rowconfigure(3, weight=1)  # la tabla crece
        main_frame.columnconfigure(0, weight=1)

        # --- Título ---
        titulo = tk.Label(main_frame, text="📋 Sistema de Inventario",
                          font=("Segoe UI", 20, "bold"),
                          bg="#f4f6f9", fg="#0078D7")
        titulo.grid(row=0, column=0, pady=10, sticky="n")

        # --- Buscador ---
        search_frame = tk.Frame(main_frame, bg="#f4f6f9")
        search_frame.grid(row=1, column=0, pady=5, sticky="ew")
        search_frame.columnconfigure(1, weight=1)

        tk.Label(search_frame, text="🔍 Buscar:", font=("Segoe UI", 11, "bold"),
                 bg="#f4f6f9").grid(row=0, column=0, padx=5)
        self.entry_buscar = tk.Entry(search_frame, font=("Segoe UI", 11))
        self.entry_buscar.grid(row=0, column=1, padx=5, sticky="ew")
        self.entry_buscar.bind("<KeyRelease>", self.buscar_producto)

        # --- Entrada de datos ---
        frame = tk.LabelFrame(main_frame, text="Agregar / Actualizar producto",
                              font=("Segoe UI", 11, "bold"),
                              bg="#f4f6f9", fg="#333")
        frame.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        for i in range(4): frame.columnconfigure(i, weight=1)

        tk.Label(frame, text="Nombre:", font=("Segoe UI", 10), bg="#f4f6f9").grid(row=0, column=0, sticky="e")
        self.entry_nombre = tk.Entry(frame, font=("Segoe UI", 10))
        self.entry_nombre.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        tk.Label(frame, text="Color:", font=("Segoe UI", 10), bg="#f4f6f9").grid(row=0, column=2, sticky="e")
        self.entry_color = tk.Entry(frame, font=("Segoe UI", 10))
        self.entry_color.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        tk.Label(frame, text="Talla:", font=("Segoe UI", 10), bg="#f4f6f9").grid(row=1, column=0, sticky="e")
        self.entry_talla = tk.Entry(frame, font=("Segoe UI", 10))
        self.entry_talla.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        tk.Label(frame, text="Stock:", font=("Segoe UI", 10), bg="#f4f6f9").grid(row=1, column=2, sticky="e")
        self.entry_stock = tk.Entry(frame, font=("Segoe UI", 10))
        self.entry_stock.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

        tk.Button(frame, text="➕ Agregar producto", font=("Segoe UI", 10, "bold"),
                  bg="#28a745", fg="white", command=self.agregar_producto)\
                  .grid(row=2, columnspan=4, pady=8)

        # --- Tabla ---
        tabla_frame = tk.Frame(main_frame, bg="#f4f6f9")
        tabla_frame.grid(row=3, column=0, padx=20, pady=10, sticky="nsew")

        self.tabla = ttk.Treeview(tabla_frame, columns=("Nombre", "Color", "Talla", "Stock"), show="headings")
        for col in ("Nombre", "Color", "Talla", "Stock"):
            self.tabla.heading(col, text=col, anchor="center")
            self.tabla.column(col, anchor="center", stretch=True)

        self.tabla.pack(fill="both", expand=True)

        # --- Botones ---
        btn_frame = tk.Frame(main_frame, bg="#f4f6f9")
        btn_frame.grid(row=4, column=0, pady=5, sticky="ew")
        for i in range(7): btn_frame.columnconfigure(i, weight=1)

        tk.Button(btn_frame, text="🗑 Eliminar", bg="#dc3545", fg="white",
                  font=("Segoe UI", 10, "bold"),
                  command=self.eliminar_producto).grid(row=0, column=0, padx=5, sticky="ew")

        tk.Button(btn_frame, text="✏️ Actualizar stock", bg="#ffc107", fg="black",
                  font=("Segoe UI", 10, "bold"),
                  command=self.actualizar_producto).grid(row=0, column=1, padx=5, sticky="ew")

        tk.Button(btn_frame, text="➕ Sumar", bg="#28a745", fg="white",
                  font=("Segoe UI", 10, "bold"),
                  command=self.sumar_stock).grid(row=0, column=2, padx=5, sticky="ew")

        tk.Button(btn_frame, text="➖ Restar", bg="#17a2b8", fg="white",
                  font=("Segoe UI", 10, "bold"),
                  command=self.restar_stock).grid(row=0, column=3, padx=5, sticky="ew")

        tk.Button(btn_frame, text="📑 Excel", bg="#0078D7", fg="white",
                  font=("Segoe UI", 10, "bold"),
                  command=self.exportar_excel).grid(row=0, column=4, padx=5, sticky="ew")

        tk.Label(btn_frame, text="Cantidad:", font=("Segoe UI", 10, "bold"),
                 bg="#f4f6f9").grid(row=0, column=5, padx=5, sticky="e")

        self.entry_cantidad = tk.Entry(btn_frame, font=("Segoe UI", 10), width=5)
        self.entry_cantidad.grid(row=0, column=6, padx=5, sticky="w")
        self.entry_cantidad.insert(0, "1")

        self.cargar_tabla()

    # --- Métodos ---
    def cargar_tabla(self, productos=None):
        self.tabla.delete(*self.tabla.get_children())
        self.tabla.tag_configure("par", background="#f9f9f9")
        self.tabla.tag_configure("impar", background="white")
        self.tabla.tag_configure("bajo_stock", background="#ffcccc")

        if productos is None:
            productos = self.inventario.productos

        for i, producto in enumerate(productos):
            tag = "par" if i % 2 == 0 else "impar"
            if producto.stock < 5:
                tag = "bajo_stock"
            self.tabla.insert("", "end",
                              values=(producto.nombre, producto.color, producto.talla, producto.stock),
                              tags=(tag,))

    def limpiar_campos(self):
        self.entry_nombre.delete(0, tk.END)
        self.entry_color.delete(0, tk.END)
        self.entry_talla.delete(0, tk.END)
        self.entry_stock.delete(0, tk.END)

    def agregar_producto(self):
        nombre = self.entry_nombre.get()
        color = self.entry_color.get()
        talla = self.entry_talla.get()
        stock = self.entry_stock.get()
        if not nombre or not color or not talla or not stock:
            messagebox.showwarning("⚠️ Error", "Todos los campos son obligatorios.")
            return
        try:
            stock = int(stock)
        except ValueError:
            messagebox.showwarning("⚠️ Error", "Stock debe ser un número entero.")
            return
        if self.inventario.agregar_producto(nombre, color, talla, stock):
            self.cargar_tabla()
            self.limpiar_campos()
            messagebox.showinfo("✅ Éxito", f"Producto '{nombre} - {color} - {talla}' agregado")
        else:
            messagebox.showwarning("⚠️ Error", "Ese producto ya existe.")

    def eliminar_producto(self):
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("⚠️ Error", "Seleccione un producto.")
            return
        valores = self.tabla.item(seleccionado[0])["values"]
        nombre, color, talla = valores[0], valores[1], valores[2]
        if self.inventario.eliminar_producto(nombre, color, talla):
            self.cargar_tabla()
            messagebox.showinfo("✅ Éxito", "Producto eliminado.")

    def actualizar_producto(self):
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("⚠️ Error", "Seleccione un producto.")
            return
        valores = self.tabla.item(seleccionado[0])["values"]
        nombre, color, talla = valores[0], valores[1], valores[2]
        try:
            nuevo_stock = int(self.entry_stock.get())
        except ValueError:
            messagebox.showwarning("⚠️ Error", "Ingrese un número válido en stock.")
            return
        if self.inventario.actualizar_producto(nombre, color, talla, nuevo_stock):
            self.cargar_tabla()
            self.limpiar_campos()
            messagebox.showinfo("✅ Éxito", "Stock actualizado.")

    def sumar_stock(self):
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("⚠️ Error", "Seleccione un producto.")
            return
        valores = self.tabla.item(seleccionado[0])["values"]
        nombre, color, talla, stock = valores[0], valores[1], valores[2], int(valores[3])
        try:
            cantidad = int(self.entry_cantidad.get())
        except ValueError:
            cantidad = 1
        nuevo_stock = stock + cantidad
        if self.inventario.actualizar_producto(nombre, color, talla, nuevo_stock):
            self.cargar_tabla()
            messagebox.showinfo("✅ Éxito", f"Stock de {nombre} ({color}, {talla}) incrementado a {nuevo_stock}")

    def restar_stock(self):
        seleccionado = self.tabla.selection()
        if not seleccionado:
            messagebox.showwarning("⚠️ Error", "Seleccione un producto.")
            return
        valores = self.tabla.item(seleccionado[0])["values"]
        nombre, color, talla, stock = valores[0], valores[1], valores[2], int(valores[3])
        try:
            cantidad = int(self.entry_cantidad.get())
        except ValueError:
            cantidad = 1
        if cantidad > stock:
            messagebox.showwarning("⚠️ Error", f"No puedes restar {cantidad}, solo hay {stock} en stock.")
            return
        nuevo_stock = stock - cantidad
        if self.inventario.actualizar_producto(nombre, color, talla, nuevo_stock):
            self.cargar_tabla()
            messagebox.showinfo("✅ Éxito", f"Stock de {nombre} ({color}, {talla}) reducido a {nuevo_stock}")

    def buscar_producto(self, event=None):
        filtro = self.entry_buscar.get().lower()
        productos_filtrados = [
            p for p in self.inventario.productos
            if filtro in p.nombre.lower() or filtro in p.color.lower() or filtro in p.talla.lower()
        ]
        self.cargar_tabla(productos_filtrados)

    def exportar_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        encabezados = ["Nombre", "Color", "Talla", "Stock"]
        ws.append(encabezados)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for col_num, col_name in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 15
        for producto in self.inventario.productos:
            ws.append([producto.nombre, producto.color, producto.talla, producto.stock])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
        archivo = "inventario.xlsx"
        wb.save(archivo)
        messagebox.showinfo("✅ Exportado", f"El inventario se guardó en {archivo}")


# --- EJECUTAR APP ---
if __name__ == "__main__":
    root = tk.Tk()
    app = InventarioApp(root)
    root.mainloop()
